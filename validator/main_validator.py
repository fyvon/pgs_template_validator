import logging
import os
import re
import urllib.request
from io import BytesIO
from urllib.error import HTTPError

from openpyxl import load_workbook

from validator.demographic import Demographic
from validator.efotrait import EFOTrait
from validator.formula import Formula
from validator.metric import Metric
from validator.performance import PerformanceMetric
from validator.publication import Publication
from validator.request.connector import DefaultConnector, ConnectorException
from validator.sample import Sample
from validator.score import Score

logger = logging.getLogger(__name__)

#---------------------#
#  General variables  #
#---------------------#

# Needed for parsing confidence intervals
insquarebrackets = re.compile(r'\[([^\)]+)\]')  # this regex might give redundant character escape warning, but they are kept for clarity
interval_format = r'^\-?\d+(e-|\.)?\d*\s\-\s\-?\d+(e-|\.)?\d*$'
inparentheses = re.compile(r'\((.*)\)')

template_columns_schema_file = os.path.join(os.path.dirname(__file__), '../templates/TemplateColumns2Models.xlsx')

# Extra fields information (not present in the Excel template schema)
metric_fields_infos = {
    'name': {'type': 'string', 'label': 'Metric - name'},
    'name_short': {'type': 'string', 'label': 'Metric - short name'},
    'type': {'type': 'string', 'label': 'Metric - type'},
    'estimate': {'type': 'float', 'label': 'Metric - Estimate value'},
    'unit': {'type': 'string', 'label': 'Metric - Unit data'},
    'se': {'type': 'float', 'label': 'Metric - Standard error value'},
    'ci': {'type': interval_format, 'label': 'Metric - Confidence interval'},
}
demographic_age_fields_infos = {
    'estimate': {'type': 'float', 'label': 'Age - Value'},
    'estimate_type': {'type': 'string', 'label': 'Age - Value type'},
    'unit': {'type': 'string', 'label': 'Age - Unit'},
    'range': {'type': interval_format, 'label': 'Age - Range'},
    'range_type': {'type': 'string', 'label': 'Age - Range type'},
    'variability': {'type': 'float', 'label': 'Age - Variablility'},
    'variability_type': {'type': 'string', 'label': 'Age - Variablility type'}
}
demographic_followup_fields_infos = {
    'estimate': {'type': 'float', 'label': 'Follow up Time - Value'},
    'estimate_type': {'type': 'string', 'label': 'Follow up Time - Value type'},
    'unit': {'type': 'string', 'label': 'Follow up Time - Unit'},
    'range': {'type': interval_format, 'label': 'Follow up Time - Range'},
    'range_type': {'type': 'string', 'label': 'Follow up Time - Range type'},
    'variability': {'type': 'float', 'label': 'Follow up Time - Variablility'},
    'variability_type': {'type': 'string', 'label': 'Follow up Time - Variablility type'}
}


class ReportError(Exception):
    """Used to interrupt a process if an identified critical validation error is detected and needs to be reported in an except clause."""


class PGSMetadataValidator():

    def __init__(self, filepath, is_remote, connector=DefaultConnector()):
        self.filepath = filepath
        self.is_remote = is_remote
        self.connector = connector
        self.parsed_publication = None
        self.parsed_scores = {}
        self.parsed_efotraits = {}
        self.parsed_samples_scores = []
        self.parsed_samples_testing = []
        self.parsed_performances = {}
        self.parsed_samplesets = []
        self.cohorts_list = []
        self.template_columns_schema_file = template_columns_schema_file
        self.table_mapschema = {}
        self.fields_infos = {}
        self.mandatory_fields = {}
        self.report = { 'error': {}, 'warning': {} }
        self.spreadsheet_names = {}
        self.scores_spreadsheet_onhold = { 'is_empty': False, 'label': '', 'error_msg': None, 'has_pgs_ids': False, 'has_testing_samples': False }


    def load_workbook_from_url(self):
        """
        Load the Excel spreadsheet into an openpyxl workbook
        > Return type: openpyxl workbooks
        """
        from google.cloud import storage
        workbook = None
        try:
            # Google cloud storage
            storage_client = storage.Client.from_service_account_json(os.environ['GS_SERVICE_ACCOUNT_SETTINGS'])
            # Fetch the required bucket
            bucket = storage_client.get_bucket(os.environ['GS_BUCKET_NAME'])
            # Fetch the data file object ("blob")
            blob = bucket.get_blob(self.filepath)
            # Download the file content
            if blob:
                data = blob.download_as_bytes()
                workbook = load_workbook(filename=BytesIO(data))
            else:
                self.report_error('General',None,'Can\'t find the uploaded file')
        except urllib.error.HTTPError as e:
            if e.code == 404:
                msg = 'The upload of the file failed'
            else:
                msg = e.reason
            self.report_error('General',None,msg)
        except Exception as e:
            self.report_error('General',None,e)
        return workbook


    #========================#
    #  Main parsing methods  #
    #========================#

    def parse_spreadsheets(self):
        """ReadCuration takes as input the location of a study metadata file"""

        self.parse_template_schema()

        loaded_spreadsheets = False
        loc_excel = self.filepath
        if loc_excel != None:
            #print("REMOTE: "+str(self.is_remote))
            if self.is_remote:
                workbook = self.load_workbook_from_url()
            else:
                workbook = load_workbook(loc_excel, data_only=True)

            if workbook:
                loaded_spreadsheets = True

                #print(str(workbook.sheetnames))

                # Check if all the spreadsheets exist in the file
                for model in self.spreadsheet_names:
                    spreadsheet_name = self.spreadsheet_names[model]
                    if not spreadsheet_name in workbook.sheetnames:
                        msg = f'The spreadsheet "{spreadsheet_name}" is missing in the Excel file.'
                        self.report_error('General',None,msg)
                        return False

                self.workbook_publication = workbook[self.spreadsheet_names['Publication']]

                self.workbook_scores = workbook[self.spreadsheet_names['Score']]

                self.workbook_samples = workbook[self.spreadsheet_names['Sample']]

                self.workbook_performances = workbook[self.spreadsheet_names['Performance']]

                self.workbook_cohorts = workbook[self.spreadsheet_names['Cohort']]

        return loaded_spreadsheets


    def parse_template_schema(self):
        """ Parse the template2model schema file. The collected and stored data will be used for the validations. """
        schema_workbook = load_workbook(self.template_columns_schema_file)
        curation_sheet = schema_workbook["Curation"]
        schema_columns = get_column_name_index(curation_sheet)
        for row_cell in curation_sheet.iter_rows(min_row=2, values_only=True):
            sheet_name = row_cell[0]
            column_name = row_cell[schema_columns['Column']]
            model_name = row_cell[schema_columns['Model']]
            field_name = row_cell[schema_columns['Field']]
            type_name = row_cell[schema_columns['Type']]
            mandatory_name = row_cell[schema_columns['Mandatory']]

            if field_name:
                if not sheet_name in self.table_mapschema:
                    self.table_mapschema[sheet_name] = {}
                self.table_mapschema[sheet_name][column_name] = field_name
                if type_name:
                    if not sheet_name in self.fields_infos:
                        self.fields_infos[sheet_name] = {}
                    column_label = trim_column_label(column_name)
                    self.fields_infos[sheet_name][field_name] = { 'type': type_name, 'label': column_label }
                if mandatory_name == 'Y':
                    if not sheet_name in self.mandatory_fields:
                        self.mandatory_fields[sheet_name] = []
                    self.mandatory_fields[sheet_name].append(field_name)

            if not model_name in self.spreadsheet_names and model_name is not None:
                self.spreadsheet_names[model_name] = sheet_name


    def parse_publication(self):
        """ Parse and validate the Publication spreadsheet. """
        spread_sheet_name = self.spreadsheet_names['Publication']
        col_names = get_column_name_index(self.workbook_publication)
        c_doi = ''
        c_PMID = ''
        row_start = 2
        row_id = row_start
        for pinfo in self.workbook_publication.iter_rows(min_row=row_start, max_row=row_start, values_only=True):
            c_doi = pinfo[col_names['doi']]
            c_doi = self.check_and_remove_whitespaces(spread_sheet_name, None, 'doi', c_doi)
            c_PMID = pinfo[1]
            c_PMID = self.check_and_remove_whitespaces(spread_sheet_name, None, 'PubMed ID', c_PMID)

        # Could have no DOI and PMID (embargoed study)
        if not c_doi and not c_PMID:
            self.report_warning(spread_sheet_name, row_id, 'No DOI or PubMed ID provided (required unless not published yet)')
            return

        # PubMed ID
        if c_PMID and c_PMID != '':
            # Removing potential .0 when PMID is converted to float
            c_PMID = str(c_PMID).removesuffix('.0')
            if not re.search(r'^\d+(?:\.0+)?$', c_PMID):
                self.report_error(spread_sheet_name,row_id,f'PubMed ID format should be only numeric or empty (found: "{c_PMID}")')

        # DOI
        if c_doi and c_doi != '':
            if not c_doi.startswith('10.'):
                self.report_error(spread_sheet_name,row_id,f'DOI format should starts with "10." or be empty but should not be an URL (found: "{c_doi}").')

        # Check in EuropePMC
        publication = Publication(c_doi, c_PMID)
        is_in_eupmc = publication.populate_from_eupmc(self.connector)
        if not is_in_eupmc:
            doi_label = ''
            if c_doi and c_doi != '':
                doi_label = f' ("{c_doi}")'
            PMID_label = ''
            if c_PMID and c_PMID != '':
                PMID_label = f' ("{c_PMID}")'
            self.report_error(spread_sheet_name,row_id,f'Can\'t find the Publication in EuropePMC: DOI{doi_label} and/or PubMed ID{PMID_label} not found')
        else:
            publication_check_report = publication.check_data(self.fields_infos[spread_sheet_name], self.mandatory_fields[spread_sheet_name])
            self.add_check_report(spread_sheet_name, row_id, publication_check_report)


    def parse_scores(self):
        """ Parse and validate the Score spreadsheet. """
        spread_sheet_name = self.spreadsheet_names['Score']
        current_schema = self.table_mapschema[spread_sheet_name]
        col_names = get_column_name_index(self.workbook_scores, row_index=2)

        row_start = 3
        trait_efo_field = 'trait_efo'
        for row_id, score_info in enumerate(self.workbook_scores.iter_rows(min_row=row_start, max_row=self.workbook_scores.max_row, values_only=True), start=row_start):
            score_name = score_info[0]
            if not score_name or score_name == '':
                break
            parsed_score = {}
            for col_name in col_names:
                val = score_info[col_names[col_name]]
                val = self.check_and_remove_whitespaces(spread_sheet_name, row_id, col_name, val)
                if col_name in current_schema and val != '' and val:
                    field = current_schema[col_name]
                    if field == trait_efo_field:
                        efo_list = [ self.check_and_remove_whitespaces(spread_sheet_name, row_id, col_name, x) for x in val.split(',') ]
                        parsed_score[field] = efo_list
                    else:
                        parsed_score[field] = val

            if trait_efo_field in parsed_score:
                for trait_efo_id in parsed_score[trait_efo_field]:
                    if not trait_efo_id in self.parsed_efotraits:
                        efo_trait = EFOTrait(trait_efo_id)
                        efo_id_found = efo_trait.populate_from_efo(self.connector)
                        if efo_id_found:
                            self.parsed_efotraits[trait_efo_id] = efo_trait
                        else:
                            self.report_error(spread_sheet_name,row_id,"Can't find a corresponding entry in EFO for '"+trait_efo_id+"'")

            # Score object and checks
            score = Score()
            score = populate_object(self.workbook_scores, score, parsed_score, self.fields_infos[spread_sheet_name])

            score_check_report = score.check_data(self.fields_infos[spread_sheet_name], self.mandatory_fields[spread_sheet_name])
            self.add_check_report(spread_sheet_name, row_id, score_check_report)

            self.parsed_scores[score_name] = score

        if not self.parsed_scores:
            self.scores_spreadsheet_onhold['is_empty'] = True
            self.scores_spreadsheet_onhold['label'] = spread_sheet_name
            self.scores_spreadsheet_onhold['error_msg'] = "No data found in this spreadsheet!"


    def parse_cohorts(self):
        """ Parse the Cohort reference spreadsheet. """
        row_start = 2
        for cohort_info in self.workbook_cohorts.iter_rows(min_row=row_start, max_row=self.workbook_cohorts.max_row, values_only=True):
            if not cohort_info or len(cohort_info) == 0 or not cohort_info[0]:
                break
            cohort_id = cohort_info[0].upper()
            cohort_id = cohort_id.strip()
            if not cohort_id in self.cohorts_list:
                self.cohorts_list.append(cohort_id)


    def cohort_to_list(self, cstring, row_id, spread_sheet_name):
        """ Check that the given cohort ID is in the Cohort spreadsheet. """
        clist = set()
        for cname in cstring.split(','):
            cname = cname.strip().upper()
            if not cname in self.cohorts_list:
                self.report_warning(spread_sheet_name,row_id,"Can't find a corresponding cohort ID in the Cohort spreadsheet for '"+str(cname)+"'")
            clist.add(cname)

        return list(clist)


    def parse_performances(self):
        """ Parse and validate the Performance Metrics spreadsheet. """
        spread_sheet_name = self.spreadsheet_names['Performance']
        current_schema = self.table_mapschema[spread_sheet_name]
        col_names = get_column_name_index(self.workbook_performances, row_index=2)

        score_names_list = self.parsed_scores.keys()

        row_start = 3
        for row_id, performance_info in enumerate(self.workbook_performances.iter_rows(min_row=row_start, max_row=self.workbook_performances.max_row, values_only=True), start=row_start):
            score_name = performance_info[0]
            if not score_name or score_name == '':
                break
            # Check that the score name is in the "Score(s)" spreadsheet. Exception if the score is an existing PGS ID.
            if self.scores_spreadsheet_onhold['is_empty'] == False:
                self.map_score_names(spread_sheet_name, row_id, score_name)
            # If the "Score(s)"" spreadsheet is empty, check that the score name is a PGS ID
            elif re.search(r'^PGS\d{6}$', score_name) and self.scores_spreadsheet_onhold['has_pgs_ids'] == False:
                self.scores_spreadsheet_onhold['has_pgs_ids'] = True

            sampleset  = performance_info[1]
            if not sampleset in self.parsed_samplesets:
                self.parsed_samplesets.append(sampleset)

            parsed_performance = {
                'score_name': score_name,
                'sampleset': sampleset
            }
            parsed_metrics = []

            for col_name in col_names:
                val = performance_info[col_names[col_name]]
                val = self.check_and_remove_whitespaces(spread_sheet_name, row_id, col_name, val)
                if (col_name in current_schema) and (val != '') and val != None:
                    field = current_schema[col_name]
                    if field.startswith('metric'):
                        for x in str(val).split(';'):
                            if x.isnumeric():
                                x = float(x)
                            try:
                                parsed_metrics.append(self.str2metric(x, row_id, spread_sheet_name, self.workbook_performances, field))
                            except ReportError as e:
                                self.report_error(spread_sheet_name, row_id, str(e))
                            except:  # Unexpected error
                                error_msg = "Error parsing the metric value '"+str(val)+"'"
                                self.report_error(spread_sheet_name, row_id, error_msg)
                    else:
                        parsed_performance[field] = val

            performance = PerformanceMetric()
            performance = populate_object(self.workbook_performances, performance, parsed_performance, self.fields_infos[spread_sheet_name])

            performance_check_report = performance.check_data(self.fields_infos[spread_sheet_name], self.mandatory_fields[spread_sheet_name])
            self.add_check_report(spread_sheet_name, row_id, performance_check_report)

            performance_id = str(parsed_performance['score_name'])+'__'+str(parsed_performance['sampleset'])
            self.parsed_performances[performance_id] = performance

            # Metrics data
            if len(parsed_metrics) > 0:
                for metric in parsed_metrics:
                    metric_check_report = metric.check_data(metric_fields_infos)
                    self.add_check_report(spread_sheet_name, row_id, metric_check_report)
            else:
                self.report_error(spread_sheet_name,row_id,"The entry is missing associated Performance Metrics data (Effect size, Classification or Other)")
        
        if not self.parsed_performances:
            self.report_error(spread_sheet_name,None,"No data found in this spreadsheet!")


    def parse_samples(self):
        """ Parse and validate the Sample spreadsheet. """
        spread_sheet_name = self.spreadsheet_names['Sample']
        current_schema = self.table_mapschema[spread_sheet_name]
        col_names = get_column_name_index(self.workbook_samples)

        samples_scores = {}
        samples_testing = {}
        # Extract data for training (GWAS + Score Development) sample
        row_start = 2
        for row_id, sample_info in enumerate(self.workbook_samples.iter_rows(min_row=row_start, max_row=self.workbook_samples.max_row, values_only=True), start=row_start):

            sample_study_type = sample_info[1]
            # Corresponds to the end of the data
            if not sample_study_type or sample_study_type == '':
                break
            sample_study_type = self.check_and_remove_whitespaces(spread_sheet_name, row_id, self.fields_infos[spread_sheet_name]['__study_stage']['label'], sample_study_type)

            if re.search('Testing',sample_study_type):
                samples_testing[row_id] = sample_info
            else:
                score_name = sample_info[0]
                if not score_name or score_name == '':
                    self.report_error(spread_sheet_name, row_id, "The column 'Associated Score Name(s)' is empty.")
                    break
                score_name = self.check_and_remove_whitespaces(spread_sheet_name, row_id, self.fields_infos[spread_sheet_name]['__score_name']['label'], score_name)
                self.map_score_names(spread_sheet_name, row_id, score_name)
                samples_scores[row_id] = sample_info

        if samples_testing and self.scores_spreadsheet_onhold['is_empty']:
            self.scores_spreadsheet_onhold['has_testing_samples'] = True

        if not samples_scores and not samples_testing:
            self.report_error(spread_sheet_name,None,"No data found in this spreadsheet!")
        else:
            self.parse_samples_scores(spread_sheet_name, current_schema, samples_scores, col_names)
            if not samples_testing:
                self.report_error(spread_sheet_name, None, "There are no 'Testing' sample entries for this study.")
            else:
                self.parse_samples_testing(spread_sheet_name, current_schema, samples_testing, col_names)

    def map_score_names(self, spreadsheet_name, row_id, scores_string: str):
        """ Attempt to map score names to those defined in the current study. """
        score_names = list(map(lambda s: s.strip(), scores_string.split(',')))
        for score_name in score_names:
            # "PGS\d{6}" score names are assumed to refer to existing PGS Catalog scores and won't be checked here
            if score_name not in self.parsed_scores and not re.match(r'^PGS\d{6}$', score_name):
                self.report_error(spreadsheet_name, row_id, f'Score name "{score_name}" can\'t be found in the Score(s) spreadsheet!')

    def parse_samples_scores(self, spread_sheet_name, current_schema, samples_scores, col_names):
        """ Parse and validate the GWAS and the Score development samples in the Sample spreadsheet. """
        samples = {}
        for row_id, sample_info in samples_scores.items():
            sample_remapped = {}
            for col_name in col_names:
                val = sample_info[col_names[col_name]]
                val = self.check_and_remove_whitespaces(spread_sheet_name, row_id, col_name, val)
                if (col_name in current_schema) and (val != '') and val != None:
                    field = current_schema[col_name]
                    if field == 'cohorts':
                        val = self.cohort_to_list(val, row_id,spread_sheet_name)
                    elif field in ['sample_age', 'followup_time']:
                        val = self.str2demographic(val, row_id, spread_sheet_name, self.workbook_samples, field, col_name)
                    sample_remapped[field] = val

            # Try to get sample data from external source
            if ('sample_number' not in sample_remapped.keys()):
                # Fetch data from GWAS Catalog
                if 'source_GWAS_catalog' in sample_remapped:
                    try:
                        gwas_study = self.get_gwas_study(sample_remapped['source_GWAS_catalog'])
                        if gwas_study:
                            for gwas_ancestry in gwas_study:
                                c_sample = sample_remapped.copy()
                                for field, val in gwas_ancestry.items():
                                    c_sample[field] = val

                                if row_id in samples:
                                    samples[row_id].append(c_sample)
                                else:
                                    samples[row_id] = [c_sample]
                        else:
                            self.report_error(spread_sheet_name, row_id, f'Can\'t fetch the GWAS information for the study {sample_remapped["source_GWAS_catalog"]}')
                    except:
                        self.report_error(spread_sheet_name, row_id, f'Can\'t fetch the GWAS information for the study {sample_remapped["source_GWAS_catalog"]}')
                else:
                    self.report_error(spread_sheet_name, row_id, f'Missing GWAS Study ID (GCST ID) to fetch the sample information')
            # Get sample data from spreadsheet
            else:
                if row_id in samples:
                    samples[row_id].append(sample_remapped)
                else:
                    samples[row_id] = [sample_remapped]

        for row_id, sample_list in samples.items():
            for sample in sample_list:
                if re.search(r'^=', str(sample['sample_number'])):
                    sample['sample_number'] = calculate_formula(self.workbook_samples,sample['sample_number'])
                try:
                    sample['sample_number'] = int(float(sample['sample_number']))
                except ValueError:
                    self.report_error(spread_sheet_name, row_id, "Can't parse the data from the column '"+self.fields_infos[spread_sheet_name]['sample_number']['label']+"': "+str(sample['sample_number']))
                    continue

                sample_object = Sample()
                sample_object = populate_object(self.workbook_samples, sample_object, sample, self.fields_infos[spread_sheet_name])

                sample_check_report = sample_object.check_data(self.fields_infos[spread_sheet_name], self.mandatory_fields[spread_sheet_name])
                self.add_check_report(spread_sheet_name, row_id, sample_check_report)

                if 'sample_age' in sample:
                    sa_check_report = sample['sample_age'].check_data(demographic_age_fields_infos)
                    self.add_check_report(spread_sheet_name, row_id, sa_check_report)
                if 'followup_time' in sample:
                    ft_check_report = sample['followup_time'].check_data(demographic_followup_fields_infos)
                    self.add_check_report(spread_sheet_name, row_id, ft_check_report)

                self.parsed_samples_scores.append(sample_object)
        
        if not self.parsed_samples_scores and self.scores_spreadsheet_onhold['is_empty'] == False:
            self.report_error(spread_sheet_name,None,"No correct Sample Score entries found in this spreadsheet (from GWAS or used in Score Development)")


    def parse_samples_testing(self, spread_sheet_name, current_schema, samples_testing, col_names):
        """ Parse and validate the testing samples in the Sample spreadsheet. """
        # Extract data Testing samples
        sample_sets_list = []

        for row_id, sample_info in samples_testing.items():
            sampleset = sample_info[2]
            sampleset = self.check_and_remove_whitespaces(spread_sheet_name, row_id, self.fields_infos[spread_sheet_name]['__sampleset']['label'], sampleset)

            if not sampleset in self.parsed_samplesets:
                    self.report_warning(spread_sheet_name, row_id, f'The Sample Set ID "{sampleset}" is not present in the \'Performance Metrics\' spreadsheet')
            if not sampleset in sample_sets_list:
                    sample_sets_list.append(sampleset)

            sample_remapped = {}
            for col_name in col_names:
                val = sample_info[col_names[col_name]]
                val = self.check_and_remove_whitespaces(spread_sheet_name, row_id, col_name, val)
                if (col_name in current_schema) and (val != '') and val != None:
                    field = current_schema[col_name]
                    if field == 'cohorts':
                        val = self.cohort_to_list(val, row_id, spread_sheet_name)
                    elif field in ['sample_age', 'followup_time']:
                        val = self.str2demographic(val, row_id, spread_sheet_name, self.workbook_samples, field, col_name)
                    sample_remapped[field] = val
            # Cohorts are not mandatory for the testing samples
            if 'cohorts' not in sample_remapped:
                self.report_warning(spread_sheet_name, row_id, "The cohorts are missing [testing sample]")

            for sample_value in ['sample_number', 'sample_cases', 'sample_controls']:
                # Check value exist for the field
                if sample_value in sample_remapped.keys():
                    if re.search(r'^=', str(sample_remapped[sample_value])):
                        # print(f'CALCULATE FORMULA FOR {sample_value}: {sample_remapped[sample_value]}')
                        sample_remapped[sample_value] = calculate_formula(self.workbook_samples,sample_remapped[sample_value])
                    try:
                        sample_remapped[sample_value] = int(float(sample_remapped[sample_value]))
                    except ValueError:
                        self.report_error(spread_sheet_name, row_id, "Can't parse the data from the column '"+self.fields_infos[spread_sheet_name][sample_value]['label']+"': "+str(sample_remapped[sample_value]))
                        continue
                else:
                    self.report_warning(spread_sheet_name, row_id, "Missing '"+self.fields_infos[spread_sheet_name][sample_value]['label']+"' value")


            sample_object = Sample()
            sample_object = populate_object(self.workbook_samples, sample_object, sample_remapped, self.fields_infos[spread_sheet_name])

            sample_check_report = sample_object.check_data(self.fields_infos[spread_sheet_name], self.mandatory_fields[spread_sheet_name])
            self.add_check_report(spread_sheet_name, row_id, sample_check_report)

            if 'sample_age' in sample_remapped:
                sa_check_report = sample_remapped['sample_age'].check_data(demographic_age_fields_infos)
                self.add_check_report(spread_sheet_name, row_id, sa_check_report)
            if 'followup_time' in sample_remapped:
                ft_check_report = sample_remapped['followup_time'].check_data(demographic_followup_fields_infos)
                self.add_check_report(spread_sheet_name, row_id, ft_check_report)

            self.parsed_samples_testing.append(sample_object)

        # Check if all the Sample Sets in the Performance Metrics spreadsheet have associated Samples
        for sampleset in self.parsed_samplesets:
            if not sampleset in sample_sets_list:
                self.report_error(spread_sheet_name, None, f'The Sample Set ID "{sampleset}" (presents in the \'Performance Metrics\' spreadsheet) has no linked samples.')

        if not self.parsed_samples_testing:
            self.report_error(spread_sheet_name,None,"No correct Sample Testing entries found in this spreadsheet")



    #=========================#
    #  Other parsing methods  #
    #=========================#

    def post_parsing_checks(self):
        """ Perform additional checks after the parsing of the spreadsheets. """

        # Score(s) spreadsheet
        if self.scores_spreadsheet_onhold['is_empty']:
            if self.scores_spreadsheet_onhold['has_pgs_ids'] == False or self.scores_spreadsheet_onhold['has_testing_samples'] == False:
                error_msg = self.scores_spreadsheet_onhold['error_msg']
                label = self.scores_spreadsheet_onhold['label']
                # Missing PGS IDs
                if self.scores_spreadsheet_onhold['has_pgs_ids'] == False:
                    error_msg += "If the study uses existing PGS Catalog Scores, they are missing in the Performance Metrics spreadsheet (e.g. PGS000001)."
                if self.scores_spreadsheet_onhold['has_testing_samples'] == False:
                    error_msg += "If the study uses existing PGS Catalog Scores, you need to provide Testing sample(s)."
                self.report_error(label,None,error_msg)


    def check_and_remove_whitespaces(self, spread_sheet_name, row_id, label, data):
        """ Check trailing spaces/tabs and remove them """
        if str(data).startswith((' ','\t')) or str(data).endswith((' ','\t')):
            label = trim_column_label(label)
            self.report_warning(spread_sheet_name, row_id, f'The content of the column \'{label}\' (i.e.: "{data}") has leading and/or trailing whitespaces.')
            data = data.strip(' \t')
        return data


    def str2metric(self, val, row_id, spread_sheet_name, wb_spreadsheet, field):
        """
        Parse and validate the metrics data from the Performance Metrics spreadsheet.
        > Parameters:
            - val: content of the cell
            - row_id: number of the current row
            - spread_sheet_name: name of the current spreadsheet
            - wb_spreadsheet: workbook instance of the current spreadsheet
            - field: corresponding field name of the current column
        > Return: instance of the Metric object
        """
        _, ftype, fname = field.split('_')

        # Find out what type of metric it is (double checking the field name)
        ftype_choices = {
            'other' : 'Other Metric',
            'beta'  : 'Effect Size',
            'class' : 'Classification Metric'
        }
        current_metric = {'type': ftype_choices[ftype]}

        # Find out if it's a common metric and stucture the information
        fname_common = {
            'OR': ('Odds Ratio', 'OR'),
            'HR': ('Hazard Ratio', 'HR'),
            'AUROC': ('Area Under the Receiver-Operating Characteristic Curve', 'AUROC'),
            'Cindex': ('Concordance Statistic', 'C-index'),
            'R2': ('Proportion of the variance explained', 'R²'),
        }
        if fname in fname_common:
            current_metric['name'] = fname_common[fname][0]
            current_metric['name_short'] = fname_common[fname][1]
        elif (ftype == 'beta') and (fname == 'other'):
            current_metric['name'] = 'Beta'
            current_metric['name_short'] = 'β'
        else:
            if '=' in val:
                fname, val = val.split('=', 1)
                current_metric['name'] = fname.strip()
            else:
                # The metric data can't be extracted. Interrupting the process with critical error.
                raise ReportError(f'Metric entry "{val}" is not in the expected format (i.e. "metrics_label = metrics_value")')

        # Parse out the confidence interval and estimate
        if type(val) == float:
            current_metric['estimate'] = val
        else:
            val = str(val)
            # matches_square = insquarebrackets.findall(val)
            # # Check if an alternative metric has been declared
            # if '=' in val:
            #     mname, val = [x.strip() for x in val.split('=')]
            #     # Check if it has short + long name
            #     matches_parentheses = inparentheses.findall(mname)
            #     if len(matches_parentheses) == 1:
            #         current_metric['name'] = mname.split('(')[0]
            #         current_metric['name_short'] = matches_parentheses[0]

            # Check if SE is reported
            matches_parentheses = inparentheses.findall(val)
            if len(matches_parentheses) == 1:
                val = val.split('(')[0].strip()
                # Check extra character/data after the parenthesis
                extra = val.strip().split(')')
                if len(extra) > 1:
                    self.report_warning(spread_sheet_name,row_id,f'Extra information detected after the parenthesis for: "{val}"')
                try:
                    current_metric['estimate'] = float(val)
                except:
                     if " " in val:
                        val, unit = val.split(" ", 1)
                        try:
                            current_metric['estimate'] = float(val)
                        except ValueError:
                            self.report_error(spread_sheet_name, row_id, f'Failed to extract metric estimate value (Expected float but found "{val}"). Is the correct separator (;) used?')
                        current_metric['unit'] = unit
                current_metric['se'] = matches_parentheses[0]
            # Extract interval
            else:
                try:
                    current_metric['estimate'] = float(val.split('[')[0])
                    # Check extra character/data after the brackets
                    extra = val.strip().split(']')
                    if len(extra) > 1:
                        # Check if second part has content
                        if (extra[1] != ''):
                            self.report_warning(spread_sheet_name,row_id,f'Extra information detected after the interval for: "{val}"')
                except:
                    self.report_error(spread_sheet_name,row_id,f'Can\'t extract the estimate value from ("{val}")')
                    current_metric['estimate'] = val
                
                matches_square = insquarebrackets.findall(val)
                if len(matches_square) == 1:
                    if re.search(interval_format, matches_square[0]):
                        try:
                            current_metric['ci'] = matches_square[0]
                            [min_ci,max_ci] = current_metric['ci'].split(' - ')
                            min_ci = float(min_ci)
                            max_ci = float(max_ci)
                            estimate = float(current_metric['estimate'])
                            # Check that the estimate is within the interval
                            if not min_ci <= estimate <= max_ci:
                                self.report_error(spread_sheet_name,row_id,f'The estimate value ("{estimate}") is not within its the confidence interval "[{min_ci} - {max_ci}]"')
                        except Exception as e:
                            self.report_error(spread_sheet_name,row_id,f'Can\'t extract the estimate value and interval from "{val}": {e}')
                    else:
                        self.report_error(spread_sheet_name,row_id,f'Confidence interval "{val}" is not in the expected format (e.g. "1.00 [0.80 - 1.20]")')

        if not 'name_short' in current_metric:
            current_metric['name_short'] = current_metric['name']

        metric_obj = Metric()
        metric_obj = populate_object(wb_spreadsheet, metric_obj, current_metric, metric_fields_infos)

        return metric_obj


    def str2demographic(self, val, row_id, spread_sheet_name, wb_spreadsheet, field, col_name):
        """
        Parse and validate the samples age and follow-up data from the Sample spreadsheet.
        > Parameters:
            - val: content of the cell
            - row_id: number of the current row 
            - spread_sheet_name: name of the current spreadsheet
            - wb_spreadsheet: workbook instance of the current spreadsheet
            - field: corresponding field name of the current column
            - col_name: full name of the column (i.e. in the header)
        > Return: instance of the Demographic object
        """
        current_demographic = {}
        if type(val) == float:
            current_demographic['estimate'] = val
        else:
            # Split by ; in case of multiple sub-fields
            l = val.split(';')
            for x in l:
                values = x.split('=')
                if len(values) == 2:
                    name = values[0].strip()
                    value = values[1].strip()
                else:
                    col = col_name.split('\n')[0]
                    col.strip(' \t\n')
                    prefix_msg = f"Wrong format in the column '{col}'"
                    if len(values) > 2:
                        prefix_msg = f"Too many values in the column '{col}'"
                    self.report_error(spread_sheet_name,row_id,f'{prefix_msg}. Format expected: "name=value_or_interval unit" (e.g. median=5.2 years).')
                    continue

                # Check if it contains a range item
                matches = insquarebrackets.findall(value)
                if len(matches) == 1:
                    if re.search(interval_format, matches[0]):
                        current_demographic['range'] = matches[0]
                    else:
                        self.report_error(spread_sheet_name,row_id,f'Data Range for the value "{value}" is not in the expected format (e.g. "1.00 [0.80 - 1.20]")')
                    current_demographic['range_type'] = name.strip()
                else:
                    if name.lower().startswith('m'):
                        current_demographic['estimate_type'] = name.strip()
                        with_units = re.match(r"([-+]?\d*\.\d+|\d+) ([a-zA-Z]+)", value, re.I)
                        if with_units:
                            items = with_units.groups()
                            current_demographic['estimate'] = items[0]
                            current_demographic['unit'] = items[1]
                        else:
                            current_demographic['estimate'] = value

                    elif name.lower().startswith('s'):
                        current_demographic['variability_type'] = name.strip()
                        with_units = re.match(r"([-+]?\d*\.\d+|\d+) ([a-zA-Z]+)", value, re.I)
                        if with_units:
                            items = with_units.groups()
                            current_demographic['variability']  = items[0]
                            current_demographic['unit'] = items[1]
                        else:
                            current_demographic['variability'] = value

        if field == 'sample_age':
            demographic_fields_infos = demographic_age_fields_infos
        else:
            demographic_fields_infos = demographic_followup_fields_infos

        demographic = Demographic()
        demographic = populate_object(wb_spreadsheet, demographic, current_demographic, demographic_fields_infos)

        return demographic


    #=================================#
    #  Error/warning reports methods  #
    #=================================#

    def report_error(self, spread_sheet_name, row_id, msg):
        """
        Store the reported error.
        - spread_sheet_name: name of the spreadsheet (e.g. Publication Information)
        - row_id: row number
        - msg: error message
        """
        if not spread_sheet_name in self.report['error']:
            self.report['error'][spread_sheet_name] = {}
        # Avoid duplicated message
        if not msg in self.report['error'][spread_sheet_name]:
            self.report['error'][spread_sheet_name][msg] = []
        # Avoid duplicated line reports
        if not row_id in self.report['error'][spread_sheet_name][msg]:
            self.report['error'][spread_sheet_name][msg].append(row_id)


    def report_warning(self, spread_sheet_name, row_id, msg):
        """
        Store the reported warning.
        - spread_sheet_name: name of the spreadsheet (e.g. Publication Information)
        - row_id: row number
        - msg: warning message
        """
        if not spread_sheet_name in self.report['warning']:
            self.report['warning'][spread_sheet_name] = {}
        # Avoid duplicated message
        if not msg in self.report['warning'][spread_sheet_name]:
            self.report['warning'][spread_sheet_name][msg] = []
        # Avoid duplicated line reports
        if not row_id in self.report['warning'][spread_sheet_name][msg]:
            self.report['warning'][spread_sheet_name][msg].append(row_id)

    def add_check_report(self, spread_sheet_name, row_id, check_report_list):
        """ Store the model check reports (errors and warnings). """
        # Error(s)
        report_error_list = check_report_list['error']
        if len(report_error_list) > 0:
            for check_report in report_error_list:
                self.report_error(spread_sheet_name,row_id,check_report)
        # Warning(s)
        report_warning_list = check_report_list['warning']
        if len(report_warning_list) > 0:
            for check_report in report_warning_list:
                self.report_warning(spread_sheet_name,row_id,check_report)


    def get_gwas_study(self, gcst_id):
        """
        Get the GWAS Study information related to the PGS sample.
        Check that all the required data is available
        > Parameter:
            - gcst_id: GWAS Study ID (e.g. GCST010127)
        > Return: list of dictionnaries (1 per ancestry)
        """
        study_data = []
        try:
            response_data = self.connector.get_gwas(gcst_id)
            if response_data:
                source_PMID = response_data['publicationInfo']['pubmedId']
                for ancestry in response_data['ancestries']:

                    if ancestry['type'] != 'initial':
                        continue

                    ancestry_data = {'source_PMID': source_PMID}
                    ancestry_data['sample_number'] = ancestry['numberOfIndividuals']

                    # ancestry_broad
                    for ancestralGroup in ancestry['ancestralGroups']:
                        if not 'ancestry_broad' in ancestry_data:
                            ancestry_data['ancestry_broad'] = ''
                        else:
                            ancestry_data['ancestry_broad'] += ','
                        ancestry_data['ancestry_broad'] += ancestralGroup['ancestralGroup']
                    # ancestry_free
                    for countryOfOrigin in ancestry['countryOfOrigin']:
                        if countryOfOrigin['countryName'] != 'NR':
                            if not 'ancestry_free' in ancestry_data:
                                ancestry_data['ancestry_free'] = ''
                            else:
                                ancestry_data['ancestry_free'] += ','
                            ancestry_data['ancestry_free'] += countryOfOrigin['countryName']

                    # ancestry_country
                    for countryOfRecruitment in ancestry['countryOfRecruitment']:
                        if countryOfRecruitment['countryName'] != 'NR':
                            if not 'ancestry_country' in ancestry_data:
                                ancestry_data['ancestry_country'] = ''
                            else:
                                ancestry_data['ancestry_country'] += ','
                            ancestry_data['ancestry_country'] += countryOfRecruitment['countryName']
                    # ancestry_additional
                    # Not found in the REST API

                    study_data.append(ancestry_data)
        except Exception as e:
            logger.debug(f'Error: can\'t fetch GWAS results for {gcst_id}: {str(e)}')

        return study_data

    def test_external_services(self):
        errors = []
        error_format_string = '%s returned an unexpected error. Check the service status.'
        try:
            self.connector.get_publication(pmid=1)
        except ConnectorException:
            errors.append(error_format_string % 'Europe PMC')

        try:
            self.connector.get_efo_trait('EFO_0001645')
        except ConnectorException:
            errors.append(error_format_string % 'Ontology Lookup Service')

        try:
            self.connector.get_gwas('GCST90132222')
        except ConnectorException:
            errors.append(error_format_string % 'GWAS Catalog')

        return errors


#=======================#
#  Independent methods  #
#=======================#

def get_column_name_index(worksheet, row_index=1):
    """ Get the list of column names and theirs indexes from a spreadsheet header.
        This is tricky sometimes as the header is spread on 2 rows for some of them. """
    col_names = {}
    col_indexes = {}
    for row in worksheet.iter_rows(min_row=1, max_row=row_index):
        for col in row:
            col_name = col.value
            if col_name:
                index = col.col_idx - 1 # 0 Based arrays in the python code
                col_indexes[index] = col_name
    for idx in col_indexes:
        col_name = col_indexes[idx]
        col_names[col_name] = idx
    return col_names


def populate_object(wb_spreadsheet, object, object_dict, object_fields):
    """ Generic method to populate a validator object. """
    for field in object_fields:
        if field.startswith('__'):
            continue
        if not hasattr(object, field) or (hasattr(object, field) and getattr(object, field) is None):
            if field in object_dict:
                if object_dict[field] is not None:
                    value = object_dict[field]
                    if re.search(r'^=', str(object_dict[field])):
                        value = calculate_formula(wb_spreadsheet,value)
                    setattr(object, field, value)
    return object


def calculate_formula(spreadsheet,data):
    """ Calculate the Excel formula if there is one """
    cell_formula = Formula(spreadsheet,data)
    calculated_value = cell_formula.formula2number()
    return calculated_value


def trim_column_label(label):
    """ Shorten the column labels in the report if it is too long. """
    column_label = label.split('\n')[0].strip(' \t')
    if len(column_label) > 40:
        column_label = column_label[:40]+'...'
    return column_label
